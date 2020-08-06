# -*- coding: utf-8 -*-
from elasticsearch import Elasticsearch
from elasticsearch.helpers import bulk
import openpyxl
import os
import uuid
from datetime import datetime
from elasticsearch_dsl import Mapping
from urllib3 import PoolManager
import json
import time

root_path = '../data'


def create_uid():
    """
    生成唯一ID
    :return:
    """
    return str(uuid.uuid1()).replace('-', '')


def create_num_uid():
    """
    生成唯一数字ID
    :return:
    """
    return int(datetime.now().strftime('%m%d%H%M%S%f'))


class ElasticObj(object):

    def __init__(self, ip, port, index_name):
        self.es = Elasticsearch([ip], port=port, timeout=60)
        self.ip = ip
        self.port = port
        self.index_name = index_name
        self.index_type = '_doc'
        self.property_list = self.get_property_list()
        self.scroll_id = None

    def bulk_index_data(self, item_list):
        """
        将数据每1000条批量写入es
        :param item_list: 写入的数据，列表
        :return:
        """
        actions = []
        for body_dict in item_list:
            # 判断ES中是否有要导入的关键字
            if self.is_property(body_dict):
                action = {
                    "_index": self.index_name,
                    "_type": self.index_type,
                    "_source": body_dict
                }
                actions.append(action)
                if len(actions) >= 1000:
                    self.__bulk(actions)
                    actions = []
        if len(actions) > 0:
            self.__bulk(actions)

    def __bulk(self, actions):
        """
        封装bulk方法，处理因为出现异常而中断的问题
        :param actions:
        :return:
        """
        success = None
        for i in range(3):
            try:
                success, _ = bulk(self.es, actions)
                print('{}向{}成功插入{}条'.format(datetime.now().strftime('%y-%m-%d %H:%M:%S'), self.index_name, success))
                break
            except Exception as e:
                print('{}请求失败,原因{}'.format(datetime.now().strftime('%y-%m-%d %H:%M:%S'), repr(e)))
                time.sleep(10)
        return success

    def get_mapping(self):
        """
        以字典的形式返回索引的mapping
        :return:
        """
        m = Mapping()
        m.update_from_es(self.index_name, using=self.es)
        return m.to_dict()

    def get_property_list(self):
        """
        返回mapping中的properies列表
        :return:
        """
        mapping_dict = self.get_mapping()
        key_list = []
        if not mapping_dict is None:
            for key in mapping_dict['_doc']['properties']:
                key_list.append(key)
        return key_list

    def is_property(self, body_dict):
        """
        判断字典的
        :param body_dict:
        :return:
        """
        ret = True
        for key in body_dict:
            if key not in self.property_list:
                print('索引{}，不存在{}'.format(self.index_name, key))
                ret = False
                break
        return ret

    def search_by_dict(self, query_dict):
        """
        根据一个字段查询，并且返回查询结果
        :param query_dict:
        :return: 列表，其中包含返回数据的字典
        """
        all_data_dict = {}
        s_dict = self.es.search(index=self.index_name, doc_type=self.index_type, body=query_dict)
        if 'error' not in s_dict:
            hits = s_dict['hits']['hits']
            for hit in hits:
                id = hit['_id']
                source = hit['_source']
                all_data_dict[id] = source
        else:
            print('查询结果错误{}'.format(str(query_dict)))
        return all_data_dict

    def count_by_dict(self, query_dict):
        """
        根据一个字典查询，并且返回符合条件的全部记录的数量
        :param query_dict:
        :return: 返回符合条件的全部记录的数量
        """
        s_dict = self.es.search(index=self.index_name, doc_type=self.index_type, body=query_dict)
        if 'error' not in s_dict:
            count = s_dict['hits']['total']
        else:
            count = 0
            print('查询结果错误{}'.format(str(query_dict)))
        return count

    def search_by_single_field(self, field, value):
        """
        根据一个字段查询，并且返回查询结果
        :param field:
        :param value:
        :return: 列表，其中包含返回数据的字典
        """
        query_dict = {
            "query": {
                "match": {
                    field: value
                }
            }
        }
        all_data_dict = self.search_by_dict(query_dict)
        return all_data_dict

    def delete_by_id(self, id):
        """
        根据_id删除记录
        :param id:
        :return:
        """
        res = self.es.delete(index=self.index_name, id=id)
        print('在索引中{}删除{}，结果{}'.format(self.index_name, id, res['result']))

    def delete_by_query(self, query_dict):
        """
        根据查询条件删除记录
        :param query_dict:
        :return:
        """
        self.es.delete_by_query(self.index_name, body=query_dict)

    def scroll_search(self, query_dict, scroll, size=1000, is_first_page=True):
        """
        深度分页查询None
        :param is_first_page:
        :param size: 每页数据量
        :param scroll: 分页生效时间,如果没有，则是已经有游标的情况
        :param query_dict:
        :return:
        """
        all_data_dict = {}
        manager = PoolManager()
        headers = {
            'Content-Type': 'application/json'
        }
        if is_first_page is True:
            # 设置分页生效时间，表示是第一次请求
            json_data = query_dict
            json_data['size'] = size
            url = '{}:{}/{}/_search?scroll={}'.format(self.ip, self.port, self.index_name, scroll)
        else:
            # 向后翻页的请求
            if self.scroll_id is not None:
                json_data = {
                    "scroll": scroll,
                    "scroll_id": self.scroll_id
                }
                url = '{}:{}/_search/scroll'.format(self.ip, self.port)
            else:
                print('scroll_id为空')
                url = None
        print('请求url{}'.format(url))
        print('请求JSON{}'.format(json_data))
        data = json.dumps(json_data)
        data = bytes(data, 'utf8')

        try:
            resp = manager.request('post', url, headers=headers, body=data)
            if str(resp.status) == '200':
                resp_dict = json.loads(resp.data.decode('utf-8'))
                # 保存获取的_scroll_id
                if '_scroll_id' in resp_dict:
                    self.scroll_id = resp_dict['_scroll_id']

                # 获取返回的数据
                if 'error' not in resp_dict:
                    hits = resp_dict['hits']['hits']
                    for hit in hits:
                        id = hit['_id']
                        source = hit['_source']
                        all_data_dict[id] = source
                else:
                    print('查询结果错误{}'.format(str(query_dict)))
            else:
                print('请求地址{}状态码不是200'.format(url))
        except Exception as e:
            print('请求地址{}失败,原因{}'.format(url, repr(e)))
        return all_data_dict


class Template(object):
    def __init__(self, ip, port, index_name):
        self.eb = ElasticObj(ip, port, index_name)
        self.root_path = '../template'
        self.index_name = index_name

    def mapping_to_excel(self):
        """
        将索引的mapping写入excel文件
        :return:
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        # 根据mapping生成模板
        key_list = self.eb.get_property_list()
        ws.append(key_list)
        wb.save(os.path.join(self.root_path, '{}.xlsx'.format(self.index_name)))
        print('保存{}的mapping完毕'.format(self.index_name))


class EsData(object):
    def __init__(self, ip, port):
        self.ip = ip
        self.port = port
        #记录运行的线程组
        self.__threads__ = []

    def convert_data_type(self, cell_type, cell_value):
        """
        根据单元格的类型，准换单元格的值
        :param cell_type:
        :param cell_value:
        :return:
        """
        # 单元格是时间格式，转换成字符串
        if cell_type == 'd':
            ret_cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
        else:
            ret_cell_value = cell_value
        return ret_cell_value

    def load(self, file_path):
        """
        从EXCEL文件中载入数据，日期类型的数据处理为字符串
        :param file_path:
        :return: 返回DataFrame数据
        """
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        all_list = []

        # 获取字典的key
        key_list = []
        first_row = list(sheet.rows)[0]
        for cell in first_row:
            if cell.value is not None:
                key_list.append(cell.value)
            else:
                print('EXCEL文件第一行存在空值的情况，不能载入数据')
                return []
        # 获取每行数据
        for i in range(2, sheet.max_row + 1):
            row_dict = {}
            row_is_all_None = True
            for j in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=i, column=j).value
                cell_type = sheet.cell(row=i, column=j).data_type
                row_dict[key_list[j - 1]] = self.convert_data_type(cell_type, cell_value)
                if cell_value is not None:
                    # 标记EXCEL的行中至少有一个单元格不是None
                    row_is_all_None = False
            # 排除EXCEL中一行都是None的数据
            if row_is_all_None is False:
                all_list.append(row_dict)
        return all_list

    def update_insert(self, index_name, data_list, key_word=None, del_old=False):
        """
        向一个索引中插入数据，如果关键字的记录存在则先删除再插入
        :param data_list: 要插入的数据
        :param index_name: 索引名称
        :param key_word: 查询数据的关键字
        :param del_old: False 不删除旧数据只插入没有的新数据, True 删除旧数据再插入新数据
        :return:
        """
        # 向ES中插入数据
        es_obj = ElasticObj(self.ip, self.port, index_name)
        add_list = []
        for row_data in data_list:
            if key_word is not None:
                # 根据一个关键字处理索引中原有的值，并且插入数据
                key_word_value = row_data[key_word]
                rs_dict = es_obj.search_by_single_field(key_word, key_word_value)
                if len(rs_dict) > 0:
                    # 索引中存在对应的值
                    if del_old is True:
                        # 删除原值,并且插入新值
                        for rs_id in rs_dict:
                            es_obj.delete_by_id(rs_id)
                        add_list.append(row_data)
                    else:
                        # 不删除原值,也不插入新值
                        pass
                else:
                    # 索引没有对应的值，则插入
                    add_list.append(row_data)
            else:
                # 不需要根据关键字处理原有的值，则直接插入数据
                add_list.append(row_data)
        # 批量插入新的数据
        es_obj.bulk_index_data(add_list)

    def del_by_single_field(self, index_name, data_list, key_word):
        """
        根据data_list中的key_word字段，在索引中删除数据
        :param index_name: 索引的名称
        :param data_list: 要删除的数据
        :param key_word: 要删除数据的关键字
        :return:
        """
        es_obj = ElasticObj(self.ip, self.port, index_name)
        for row_data in data_list:
            key_word_value = row_data[key_word]
            rs_dict = es_obj.search_by_single_field(key_word, key_word_value)
            if len(rs_dict) > 0:
                for rs_id in rs_dict:
                    es_obj.delete_by_id(rs_id)

    def update_insert_from_excel(self, file_path, key_word=None, del_old=False, str_key_list=[], int_key_list=[]):
        """
        从EXCEL中读取数据并且插入ES
        :param str_key_list: 列表中的关键字，将被一个字符串替换
        :param int_key_list: 列表中的关键字，将被一个随机数字替换
        :param key_word: 查询数据的关键字
        :param del_old: False 不删除旧数据只插入没有的新数据, True 删除旧数据再插入新数据
        :param file_path:
        :return:
        """
        # 根据文件名称获取索引名称
        index_name = os.path.basename(file_path).replace('.xlsx', '')
        # 从excel中载入数据
        all_list = self.load(file_path)
        for row in all_list:
            for key in str_key_list:
                row[key] = create_uid()
            for key in int_key_list:
                row[key] = create_num_uid()
        self.update_insert(index_name, all_list, key_word, del_old)

    def del_from_excel(self, file_path, key_word):
        """
        从EXCEL中根据关键字删除记录
        :param file_path:
        :param key_word: 查询数据的关键字
        :return:
        """
        # 根据文件名称获取索引名称
        index_name = os.path.basename(file_path).replace('.xlsx', '')
        # 从excel中载入数据
        all_list = self.load(file_path)
        self.del_by_single_field(index_name, all_list, key_word)

    def big_insert_from_excel(self, file_path, times=1, str_key_list=[], int_key_list=[]):
        """
        插入大数据量的数据使用
        :param thread_num: 开启的线程数
        :param str_key_list: 列表中的关键字，将被一个字符串替换
        :param int_key_list: 列表中的关键字，将被一个随机数字替换
        :param file_path:
        :param times 数据插入的倍数
        :return:
        """
        index_name = os.path.basename(file_path).replace('.xlsx', '')
        # 从excel中载入数据
        all_list = self.load(file_path)
        # 向ES中插入数据
        for n in range(0, times):
            for row in all_list:
                for key in str_key_list:
                    row[key] = create_uid()
                for key in int_key_list:
                    row[key] = create_num_uid()
            self.update_insert(index_name, all_list)
            print('完成{}倍数据插入'.format(n+1))


def dump(src_ip, src_port, src_index, dest_ip, dest_port, dest_index):
    """
    将一个索引中的记录复制到另外一个索引
    :param src_ip: 源IP
    :param src_port: 源端口
    :param src_index: 源索引
    :param dest_ip: 目的IP
    :param dest_port: 目的端口
    :param dest_index: 目的索引
    :return:
    """
    print('开始从{}:{}/{}向{}:{}/{}导入数据'.format(src_ip, src_port, src_index, dest_ip, dest_port, dest_index))
    src_es_obj = ElasticObj(src_ip, src_port, src_index)
    dest_es_obj = ElasticObj(dest_ip, dest_port, dest_index)

    # 查询所有的数据
    query_dict = {
        "query": {
            "match_all": {}
        }
    }

    src_data_dict = src_es_obj.scroll_search(query_dict, scroll='5m', size=10000, is_first_page=True)
    while len(src_data_dict) > 0:
        dest_es_obj.bulk_index_data(src_data_dict.values())
        src_data_dict = src_es_obj.scroll_search(query_dict, scroll='5m', is_first_page=False)
    print('导入完毕')


if __name__ == '__main__':
    # 初始化ES
    es = EsData('10.108.2.181', 9200)

    # 利用EXCEL更新ES数据。EXCEL文件名既是索引名称，EXCEL列第一行的值既是字段名称
    # 插入数据时，根据指定的关键字更新旧数据
    es.update_insert_from_excel(
        os.path.join(root_path, 'anji_ajzyapp_data_operationsevents_anji_test.xlsx'), 'id', del_old=True)
    # 插入数据时，不更新旧数据，只增加新数据。并且设置某个字段的值是唯一字符串
    # es.update_insert_from_excel(
    #     os.path.join(root_path, 'anji_ajzyapp_data_allevents_customer_test.xlsx'), str_key_list=['id'])

    # 根据关键字删除数据
    # es.del_from_excel(
    #     os.path.join(root_path, 'anji_ajzyapp_data_operationsevents_anji_test.xlsx'), 'id')

    # # 将索引的mapping写入EXCEL，生成数据维护的模板
    # index_list = ['anji_ajzyapp_data_allevents_anji_test',
    #               'anji_ajzyapp_data_allevents_customer_test',
    #               'anji_ajzyapp_data_allevents_shipment_test',
    #               'anji_ajzyapp_data_allorders_test',
    #               'anji_ajzyapp_data_allshipments_test',
    #               'anji_ajzyapp_data_operationsevents_anji_test',
    #               'anji_ajzyapp_data_operationsevents_customer_test',
    #               'anji_ajzyapp_data_operationsevents_shipment_test',
    #               'anji_ajzyapp_rpt_indexdetailday_test']
    # # index_list = ['anji_ajzyapp_data_allorders_dev']
    # for index in index_list:
    #     t = Template('10.108.2.181', 9200, index)
    #     t.mapping_to_excel()

    # 把EXCEL中的数据读出来，并且成倍插入ES。用于制造性能测试用的数据
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_allevents_shipment_test.xlsx'),
    #                          times=25900,
    #                          str_key_list=['aj_big_order_no', 'customer_order_no', 'order_generate_no', 'id',
    #                                        'aj_small_order_no','aj_order_no','id','shipment_no','vin', 'vin_suffix'],
    #                          int_key_list=['shipment_id'])
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_allevents_anji_test.xlsx'),
    #                          times=10000,
    #                          str_key_list=['aj_order_no', 'customer_order_no', 'order_generate_no',
    #                                        'id', 'vin', 'vin_suffix'])
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_allevents_customer_test.xlsx'),
    #                          times=4667,
    #                          str_key_list=['aj_big_order_no', 'customer_order_no', 'order_generate_no',
    #                                        'id', 'vin', 'vin_suffix'])
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_allshipments_test.xlsx'),
    #                          times=10000,
    #                          str_key_list=['aj_big_order_no', 'customer_order_no', 'order_generate_no',
    #                                        'order_no', 'shipment_no', 'vin', 'vin_suffix'],
    #                          int_key_list=['shipment_id'])
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_allorders_test.xlsx'),
    #                          times=5000,
    #                          str_key_list=['aj_big_order_no', 'customer_order_no', 'order_generate_no',
    #                                        'vin', 'vin_suffix'])
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_operationsevents_shipment_test.xlsx'),
    #                          times=1420,
    #                          str_key_list=['aj_big_order_no', 'customer_order_no', 'order_generate_no', 'id',
    #                                        'aj_small_order_no','aj_order_no','id','shipment_no','vin', 'vin_suffix'],
    #                          int_key_list=['shipment_id'])
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_operationsevents_anji_test.xlsx'),
    #                          times=1280,
    #                          str_key_list=['aj_order_no', 'customer_order_no', 'order_generate_no',
    #                                        'id', 'vin', 'vin_suffix'])
    # es.big_insert_from_excel(os.path.join(root_path, 'anji_ajzyapp_data_operationsevents_customer_test.xlsx'),
    #                          times=1010,
    #                          str_key_list=['aj_big_order_no', 'customer_order_no', 'order_generate_no',
    #                                        'id', 'vin', 'vin_suffix'])
    # 将一个索引中的数据复制到另外一个索引中
    # dump('10.108.141.245', 9200, 'anji_ajzyapp_rpt_indexdetailday_uat',
    #      '10.108.2.181', 9200, 'anji_ajzyapp_rpt_indexdetailday_test')
